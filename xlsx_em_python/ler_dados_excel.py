from openpyxl import load_workbook

wb=load_workbook("palavras.xlsx")

ws = wb.active
primeira_coluna = ws['B']
palavras = []

for i in range(len(primeira_coluna)):
	palavras.append(primeira_coluna[i].value) 

print("Palavras que estão no arquivo excel: ")
for i in range(len(palavras)):
	print(palavras[i])

