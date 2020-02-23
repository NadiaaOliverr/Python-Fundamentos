import openpyxl

def cria_planilha(titulo_da_planilha, produtos, precos):
    wb = openpyxl.Workbook()

    planilha_produtos = wb.active

    planilha_produtos.title = titulo_da_planilha
    planilha_produtos.column_dimensions['A'].width = 15
    planilha_produtos.column_dimensions['B'].width = 10

    planilha_produtos['A1'] = "Produto"
    planilha_produtos['B1'] = "Preço"

    for i in range(len(produtos)):
        produto = planilha_produtos.cell(row=i+2, column=1)
        produto.value = produtos[i]

        preco = planilha_produtos.cell(row=i+2, column=2)
        preco.style = 'Currency'
        preco.number_format = 'R$ #,##0.00'
        preco.value = precos[i]

    wb.save("lista_de_produtos.xlsx")


produtos = ['SSD 120GB', 'HD 1TB', 'M.2 960GB', 'Memória 8GB']
precos = [160.50, 230.10, 899.90, 179.90]

cria_planilha('Lista de Produtos', produtos, precos)