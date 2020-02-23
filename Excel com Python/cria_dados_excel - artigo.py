import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.Workbook()

planilha_produtos = wb.active

planilha_produtos.title = "Produtos"

titulo_coluna_produtos = "Produto"
titulo_coluna_preco = "Preço"


planilha_produtos['A1'].fill = PatternFill(start_color='32A7E5',end_color='32A7E5', fill_type = "solid")
planilha_produtos['B1'].fill = PatternFill(start_color='32A7E5',end_color='32A7E5', fill_type = "solid")

planilha_produtos['A1'] = titulo_coluna_produtos
planilha_produtos['B1'] = titulo_coluna_preco


#se fossem mais células, poderiamos iterar sobre as células


produtos = ['SSD 120GB', 'HD 1TB', 'M.2 960GB', 'Memória 8GB DDR4']
precos = [160.50, 230.10, 899.90, 179.90]

# planilha_produtos.style = 'Currency'

# planilha_produtos.number_format = '#,##0.00 R$' 

planilha_produtos.column_dimensions['A'].width = 20
planilha_produtos.column_dimensions['B'].width = 20


for i in range(len(produtos)):
    c1 = planilha_produtos.cell(row=i+2, column=1)
    c1.value = produtos[i]

    c2 = planilha_produtos.cell(row=i+2, column=2)
    c2.style = 'Currency'
    c2.number_format = 'R$ #,##0.00'
    c2.value = precos[i]

planilha_produtos['A6'] = "Total"
total = planilha_produtos.cell(row=6, column=2)
total.style = 'Currency'
total.number_format = 'R$ #,##0.00'
total.value = sum(precos)


wb.save("produtos.xlsx")
